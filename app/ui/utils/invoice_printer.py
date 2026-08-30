# -*- coding: utf-8 -*-
"""ابزارهای چاپ و خروجی فاکتور"""

import io
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QFileDialog, QMessageBox, QWidget
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class InvoicePrinter:
    """چاپ و خروجی فاکتور در فرمت‌های مختلف"""

    @staticmethod
    def print_pdf(
        invoice_data: Dict,
        items: List[Dict],
        parent: Optional[QWidget] = None
    ) -> bool:
        """چاپ فاکتور به صورت PDF"""
        if parent:
            file_path, _ = QFileDialog.getSaveFileName(
                parent,
                "ذخیره فاکتور PDF",
                f"فاکتور_{invoice_data.get('id', '')}_{datetime.now().strftime('%Y%m%d')}.pdf",
                "PDF Files (*.pdf)"
            )
        else:
            file_path = f"فاکتور_{invoice_data.get('id', '')}_{datetime.now().strftime('%Y%m%d')}.pdf"

        if not file_path:
            return False

        if not file_path.endswith('.pdf'):
            file_path += '.pdf'

        try:
            InvoicePrinter._create_pdf(file_path, invoice_data, items)
            if parent:
                QMessageBox.information(
                    parent,
                    "موفق",
                    f"فاکتور با موفقیت در مسیر زیر ذخیره شد:\n{file_path}"
                )
            return True
        except Exception as e:
            if parent:
                QMessageBox.critical(
                    parent,
                    "خطا",
                    f"خطا در ایجاد PDF:\n{str(e)}"
                )
            return False

    @staticmethod
    def _create_pdf(file_path: str, invoice_data: Dict, items: List[Dict]) -> None:
        """ایجاد فایل PDF"""
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm,
        )

        styles = getSampleStyleSheet()
        story = []

        # استایل‌های سفارشی
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=20,
            alignment=1,
            spaceAfter=12,
            textColor=colors.HexColor('#4b2fa8')
        )

        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=1,
            spaceAfter=6,
        )

        normal_style = ParagraphStyle(
            'NormalStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=0,
            fontName='Helvetica',
        )

        story.append(Paragraph("فاکتور فروش", title_style))
        story.append(Spacer(1, 0.3*cm))

        info_data = [
            [Paragraph(f"شماره فاکتور: {invoice_data.get('id', '')}", normal_style),
             Paragraph(f"تاریخ: {invoice_data.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M'))}", normal_style)],
            [Paragraph(f"مشتری: {invoice_data.get('customer', 'مشتری متفرقه')}", normal_style),
             Paragraph(f"فروشنده: {invoice_data.get('seller', 'سیستم')}", normal_style)],
        ]

        info_table = Table(info_data, colWidths=[8*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.5*cm))

        # جدول اقلام
        table_data = []
        headers = ['ردیف', 'نام کالا', 'تعداد', 'قیمت واحد', 'جمع']
        table_data.append([Paragraph(h, header_style) for h in headers])

        for idx, item in enumerate(items, 1):
            row = [
                Paragraph(str(idx), normal_style),
                Paragraph(item.get('name', ''), normal_style),
                Paragraph(str(item.get('quantity', 0)), normal_style),
                Paragraph(f"{item.get('unit_price', 0):,.0f}", normal_style),
                Paragraph(f"{item.get('line_total', item.get('quantity', 0) * item.get('unit_price', 0)):,.0f}", normal_style),
            ]
            table_data.append(row)

        col_widths = [1.5*cm, 6*cm, 2*cm, 3*cm, 3.5*cm]
        items_table = Table(table_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0ecff')),
            ('FONTWEIGHT', (0, 0), (-1, 0), 'BOLD'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4b2fa8')),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 0.5*cm))

        # جمع‌بندی
        subtotal = sum(item.get('quantity', 0) * item.get('unit_price', 0) for item in items)
        discount = invoice_data.get('discount', 0)
        total = invoice_data.get('total', subtotal - discount)

        summary_data = [
            ['جمع کل:', f"{subtotal:,.0f}"],
            ['تخفیف:', f"{discount:,.0f}"],
            ['مبلغ قابل پرداخت:', f"{max(total, 0):,.0f}"],
        ]

        summary_table = Table(summary_data, colWidths=[5*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTWEIGHT', (0, -1), (1, -1), 'BOLD'),
            ('TEXTCOLOR', (0, -1), (1, -1), colors.HexColor('#d32f2f')),
            ('BACKGROUND', (0, -1), (1, -1), colors.HexColor('#fff5f5')),
        ]))
        story.append(summary_table)

        doc.build(story)

    @staticmethod
    def print_excel(
        invoice_data: Dict,
        items: List[Dict],
        parent: Optional[QWidget] = None
    ) -> bool:
        """چاپ فاکتور به صورت Excel"""
        if parent:
            file_path, _ = QFileDialog.getSaveFileName(
                parent,
                "ذخیره فاکتور Excel",
                f"فاکتور_{invoice_data.get('id', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
        else:
            file_path = f"فاکتور_{invoice_data.get('id', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        if not file_path:
            return False

        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            InvoicePrinter._create_excel(file_path, invoice_data, items)
            if parent:
                QMessageBox.information(
                    parent,
                    "موفق",
                    f"فاکتور با موفقیت در مسیر زیر ذخیره شد:\n{file_path}"
                )
            return True
        except Exception as e:
            if parent:
                QMessageBox.critical(
                    parent,
                    "خطا",
                    f"خطا در ایجاد Excel:\n{str(e)}"
                )
            return False

    @staticmethod
    def _create_excel(file_path: str, invoice_data: Dict, items: List[Dict]) -> None:
        """ایجاد فایل Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "فاکتور فروش"

        # استایل‌ها
        title_font = Font(name='B Nazanin', size=16, bold=True, color='4B2FA8')
        header_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='B Nazanin', size=11)
        bold_font = Font(name='B Nazanin', size=12, bold=True)
        total_font = Font(name='B Nazanin', size=14, bold=True, color='D32F2F')

        header_fill = PatternFill(start_color='4B2FA8', end_color='4B2FA8', fill_type='solid')
        total_fill = PatternFill(start_color='FFF5F5', end_color='FFF5F5', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        ws.merge_cells('A1:E1')
        cell = ws.cell(row=1, column=1, value="فاکتور فروش")
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

        row = 3
        info = [
            (f"شماره فاکتور: {invoice_data.get('id', '')}", f"تاریخ: {invoice_data.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M'))}"),
            (f"مشتری: {invoice_data.get('customer', 'مشتری متفرقه')}", f"فروشنده: {invoice_data.get('seller', 'سیستم')}"),
        ]
        
        for info_row in info:
            ws.cell(row=row, column=1, value=info_row[0]).font = normal_font
            ws.cell(row=row, column=3, value=info_row[1]).font = normal_font
            row += 1

        row += 1

        headers = ['ردیف', 'نام کالا', 'تعداد', 'قیمت واحد', 'جمع']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for idx, item in enumerate(items, 1):
            row += 1
            values = [
                idx,
                item.get('name', ''),
                item.get('quantity', 0),
                item.get('unit_price', 0),
                item.get('line_total', item.get('quantity', 0) * item.get('unit_price', 0))
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        row += 2
        subtotal = sum(item.get('quantity', 0) * item.get('unit_price', 0) for item in items)
        discount = invoice_data.get('discount', 0)
        total = invoice_data.get('total', subtotal - discount)

        summary_data = [
            ('جمع کل:', subtotal),
            ('تخفیف:', discount),
            ('مبلغ قابل پرداخت:', max(total, 0))
        ]

        for label, value in summary_data:
            row += 1
            ws.cell(row=row, column=3, value=label).font = bold_font
            ws.cell(row=row, column=4, value=value).font = bold_font
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='left')

        last_row = row
        ws.cell(row=last_row, column=3).font = total_font
        ws.cell(row=last_row, column=4).font = total_font
        ws.cell(row=last_row, column=3).fill = total_fill
        ws.cell(row=last_row, column=4).fill = total_fill

        wb.save(file_path)

    @staticmethod
    def print_preview(
        invoice_data: Dict,
        items: List[Dict],
        parent: Optional[QWidget] = None
    ) -> None:
        """نمایش پیش‌نمایش فاکتور در یک دیالوگ"""
        from PySide6.QtWidgets import QDialog, QTextBrowser, QPushButton, QHBoxLayout, QVBoxLayout

        def _money(value: float) -> str:
            return f"{value:,.0f}"

        dialog = QDialog(parent)
        dialog.setWindowTitle(f"پیش‌نمایش فاکتور #{invoice_data.get('id', '')}")
        dialog.resize(700, 600)
        dialog.setLayoutDirection(Qt.RightToLeft)

        preview = QTextBrowser()
        
        rows = "".join(
            f"""
            <tr>
                <td>{idx}</td>
                <td>{item.get('name', '')}</td>
                <td align='center'>{item.get('quantity', 0)}</td>
                <td align='center'>{_money(item.get('unit_price', 0))}</td>
                <td align='center'>{_money(item.get('line_total', item.get('quantity', 0) * item.get('unit_price', 0)))}</td>
            </tr>
            """
            for idx, item in enumerate(items, 1)
        )

        subtotal = sum(item.get('quantity', 0) * item.get('unit_price', 0) for item in items)
        discount = invoice_data.get('discount', 0)
        total = invoice_data.get('total', subtotal - discount)

        html = f"""
        <html dir="rtl">
        <head>
            <style>
                body {{ font-family: Tahoma, sans-serif; padding: 20px; }}
                h2 {{ color: #4b2fa8; text-align: center; }}
                .info {{ background: #f6f4ff; padding: 10px; border-radius: 8px; margin: 10px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th {{ background: #4b2fa8; color: white; padding: 8px; }}
                td {{ padding: 6px; border: 1px solid #ddd; }}
                .total {{ font-size: 16px; font-weight: bold; color: #d32f2f; background: #fff5f5; padding: 10px; border-radius: 8px; margin-top: 10px; }}
                .summary {{ font-weight: bold; }}
            </style>
        </head>
        <body>
            <h2>فاکتور فروش</h2>
            
            <div class="info">
                <p><strong>شماره فاکتور:</strong> {invoice_data.get('id', '')} &nbsp;&nbsp;|&nbsp;&nbsp;
                   <strong>تاریخ:</strong> {invoice_data.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M'))}</p>
                <p><strong>مشتری:</strong> {invoice_data.get('customer', 'مشتری متفرقه')} &nbsp;&nbsp;|&nbsp;&nbsp;
                   <strong>فروشنده:</strong> {invoice_data.get('seller', 'سیستم')}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>ردیف</th>
                        <th>نام کالا</th>
                        <th>تعداد</th>
                        <th>قیمت واحد</th>
                        <th>جمع</th>
                    </tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
            
            <div class="summary">
                <p><strong>جمع کل:</strong> {_money(subtotal)}</p>
                <p><strong>تخفیف:</strong> {_money(discount)}</p>
                <div class="total">
                    <strong>مبلغ قابل پرداخت:</strong> {_money(max(total, 0))}
                </div>
            </div>
        </body>
        </html>
        """
        
        preview.setHtml(html)

        btn_layout = QHBoxLayout()
        
        btn_pdf = QPushButton("📄 PDF")
        btn_pdf.clicked.connect(lambda: InvoicePrinter.print_pdf(invoice_data, items, dialog))
        
        btn_excel = QPushButton("📊 Excel")
        btn_excel.clicked.connect(lambda: InvoicePrinter.print_excel(invoice_data, items, dialog))
        
        btn_close = QPushButton("بستن")
        btn_close.clicked.connect(dialog.accept)
        
        btn_layout.addWidget(btn_pdf)
        btn_layout.addWidget(btn_excel)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)

        layout = QVBoxLayout(dialog)
        layout.addWidget(preview)
        layout.addLayout(btn_layout)

        dialog.exec()