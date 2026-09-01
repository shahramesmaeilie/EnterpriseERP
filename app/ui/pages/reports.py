# -*- coding: utf-8 -*-
"""صفحه گزارش‌گیری پیشرفته با نمودار و خروجی Excel/PDF"""

from __future__ import annotations

import sqlite3
import os
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QTimer, QDate, Signal
from PySide6.QtGui import QColor, QPainter
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QDateEdit, QGroupBox, QFrame, QMessageBox, QTabWidget,
    QSpinBox, QCheckBox, QProgressBar, QFileDialog,
    QSplitter, QScrollArea, QGridLayout, QSizePolicy,
    QLineEdit,
)
from PySide6.QtCharts import (
    QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis,
    QValueAxis, QPieSeries, QPieSlice, QLineSeries,
    QCategoryAxis, QLegend, QHorizontalBarSeries
)

import jdatetime

from app.ui.widgets import JalaliDateEdit

try:
    from app.database.migrations import DB_PATH
except Exception:
    DB_PATH = Path(__file__).resolve().parents[3] / "enterprise.db"


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _money(value: float) -> str:
    return f"{value:,.0f}"


def _to_jalali(date_str: str) -> str:
    """تبدیل تاریخ میلادی به شمسی"""
    if not date_str:
        return ""
    try:
        if isinstance(date_str, str):
            date_part = date_str.split(' ')[0]
            if 'T' in date_part:
                date_part = date_part.split('T')[0]
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            jalali_date = jdatetime.datetime.fromgregorian(datetime=dt)
            return jalali_date.strftime("%Y/%m/%d")
        elif isinstance(date_str, datetime):
            jalali_date = jdatetime.datetime.fromgregorian(datetime=date_str)
            return jalali_date.strftime("%Y/%m/%d")
        else:
            return str(date_str)
    except Exception:
        return str(date_str)


def _to_jalali_datetime(date_str: str) -> str:
    """تبدیل تاریخ و زمان میلادی به شمسی"""
    if not date_str:
        return ""
    try:
        if isinstance(date_str, str):
            parts = date_str.split(' ')
            date_part = parts[0]
            time_part = parts[1] if len(parts) > 1 else ""
            if 'T' in date_part:
                date_part = date_part.split('T')[0]
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            jalali_date = jdatetime.datetime.fromgregorian(datetime=dt)
            if time_part:
                return f"{jalali_date.strftime('%Y/%m/%d')} {time_part[:5]}"
            return jalali_date.strftime("%Y/%m/%d %H:%M")
        else:
            return str(date_str)
    except Exception:
        return str(date_str)


def _jalali_to_gregorian_string(jalali_date: jdatetime.date) -> str:
    """تبدیل تاریخ شمسی به رشته میلادی با مدیریت خطا"""
    if not jalali_date:
        return "2000-01-01"
    try:
        gregorian = jalali_date.togregorian()
        return f"{gregorian.year}-{gregorian.month:02d}-{gregorian.day:02d}"
    except AttributeError:
        try:
            gregorian = jalali_date.to_gregorian()
            return f"{gregorian.year}-{gregorian.month:02d}-{gregorian.day:02d}"
        except:
            try:
                gregorian = jdatetime.date.togregorian(jalali_date)
                return f"{gregorian.year}-{gregorian.month:02d}-{gregorian.day:02d}"
            except:
                year = jalali_date.year + 621
                month = jalali_date.month
                day = jalali_date.day
                if month > 6:
                    year += 1
                    month = month - 6
                return f"{year}-{month:02d}-{day:02d}"


def _open_file(file_path: str) -> None:
    """باز کردن فایل با برنامه پیش‌فرض سیستم"""
    try:
        if sys.platform == 'win32':
            os.startfile(file_path)
        elif sys.platform == 'darwin':
            subprocess.run(['open', file_path])
        else:
            subprocess.run(['xdg-open', file_path])
    except Exception as e:
        print(f"Error opening file: {e}")


class ReportsPage(QWidget):
    """صفحه گزارش‌گیری با نمودارهای تعاملی و خروجی"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self._setup_ui()
        self._apply_styles()
        self._load_initial_data()

    def _setup_ui(self):
        """راه‌اندازی رابط کاربری"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        # عنوان صفحه
        title_layout = QHBoxLayout()
        title = QLabel("📊 گزارش‌ها و آمار", objectName="pageTitle")
        title_layout.addWidget(title)
        title_layout.addStretch()

        # دکمه‌های اکشن
        self.btn_refresh = QPushButton("🔄 بروزرسانی")
        self.btn_refresh.setObjectName("btnPrimary")
        self.btn_refresh.clicked.connect(self._refresh_all)
        title_layout.addWidget(self.btn_refresh)

        self.btn_export = QPushButton("📤 خروجی Excel")
        self.btn_export.setObjectName("btnPrimary")
        self.btn_export.clicked.connect(self._export_excel)
        title_layout.addWidget(self.btn_export)

        self.btn_export_pdf = QPushButton("📄 خروجی PDF")
        self.btn_export_pdf.setObjectName("btnPrimary")
        self.btn_export_pdf.clicked.connect(self._export_pdf)
        title_layout.addWidget(self.btn_export_pdf)

        main_layout.addLayout(title_layout)

        # تب‌های گزارش
        self.tabs = QTabWidget()
        self.tabs.setObjectName("reportTabs")

        # تب‌های مختلف
        self.tabs.addTab(self._create_dashboard_tab(), "📊 داشبورد")
        self.tabs.addTab(self._create_sales_tab(), "💰 فروش")
        self.tabs.addTab(self._create_purchases_tab(), "🛒 خرید")
        self.tabs.addTab(self._create_inventory_tab(), "📦 انبار")
        self.tabs.addTab(self._create_products_tab(), "📋 محصولات")
        self.tabs.addTab(self._create_customers_tab(), "👥 مشتریان")

        main_layout.addWidget(self.tabs, 1)

    def _create_dashboard_tab(self) -> QWidget:
        """تب داشبورد اصلی با کارت‌های آماری و نمودارها"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)

        # کارت‌های آماری
        stats_layout = QGridLayout()
        stats_layout.setSpacing(12)

        stats = [
            ("💰 کل فروش", self._get_total_sales(), "#5b3ec8"),
            ("🛒 تعداد خرید", self._get_total_purchases(), "#2e7d32"),
            ("📦 موجودی کالا", self._get_total_products(), "#c62828"),
            ("👥 تعداد مشتریان", self._get_total_customers(), "#e65100"),
        ]

        for idx, (label, value, color) in enumerate(stats):
            card = self._create_stat_card(label, value, color)
            stats_layout.addWidget(card, idx // 2, idx % 2)

        layout.addLayout(stats_layout)

        # نمودارها
        charts_layout = QHBoxLayout()
        charts_layout.setSpacing(12)

        sales_chart = self._create_monthly_sales_chart()
        charts_layout.addWidget(sales_chart, 2)

        products_chart = self._create_products_distribution_chart()
        charts_layout.addWidget(products_chart, 1)

        layout.addLayout(charts_layout)

        return tab

    def _create_stat_card(self, title: str, value: str, color: str) -> QFrame:
        """ساخت کارت آماری"""
        card = QFrame()
        card.setObjectName("statCard")
        card.setStyleSheet(f"""
            #statCard {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #f8f6ff);
                border: 2px solid {color}22;
                border-radius: 16px;
                padding: 16px;
                min-height: 80px;
            }}
            #statCard:hover {{
                border-color: {color};
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setSpacing(4)

        lbl_title = QLabel(title)
        lbl_title.setStyleSheet(f"color: {color}; font-size: 13px; font-weight: 600;")

        lbl_value = QLabel(str(value))
        lbl_value.setStyleSheet("color: #1a1a2e; font-size: 24px; font-weight: bold;")

        layout.addWidget(lbl_title)
        layout.addWidget(lbl_value)

        return card

    def _create_monthly_sales_chart(self) -> QWidget:
        """نمودار فروش ماهانه"""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)

        label = QLabel("📈 فروش ماهانه (میلیون ریال)")
        label.setStyleSheet("font-weight: bold; color: #4a3a6a; font-size: 14px;")
        layout.addWidget(label)

        chart_view = QChartView()
        chart_view.setRenderHint(QPainter.Antialiasing)

        chart = QChart()
        chart.setTitle("")
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundVisible(False)
        chart.setTheme(QChart.ChartThemeLight)

        sales_data = self._get_monthly_sales()

        series = QLineSeries()
        series.setName("فروش")
        series.setColor(QColor("#5b3ec8"))
        series.setPointsVisible(True)

        for i, (month, amount) in enumerate(sales_data):
            series.append(i, amount / 1_000_000)

        chart.addSeries(series)

        axis_x = QCategoryAxis()
        axis_x.setLabelsAngle(-45)
        for i, (month, _) in enumerate(sales_data):
            axis_x.append(month, i)
        chart.addAxis(axis_x, Qt.AlignBottom)
        series.attachAxis(axis_x)

        axis_y = QValueAxis()
        axis_y.setLabelFormat("%.0f")
        axis_y.setTitleText("میلیون ریال")
        chart.addAxis(axis_y, Qt.AlignLeft)
        series.attachAxis(axis_y)

        chart_view.setChart(chart)
        layout.addWidget(chart_view)

        return container

    def _create_products_distribution_chart(self) -> QWidget:
        """نمودار توزیع محصولات"""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)

        label = QLabel("📊 توزیع موجودی محصولات")
        label.setStyleSheet("font-weight: bold; color: #4a3a6a; font-size: 14px;")
        layout.addWidget(label)

        chart_view = QChartView()
        chart_view.setRenderHint(QPainter.Antialiasing)

        chart = QChart()
        chart.setTitle("")
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundVisible(False)

        series = QPieSeries()

        products_data = self._get_products_distribution()

        for name, count in products_data[:6]:
            slice_data = QPieSlice(name, count)
            slice_data.setLabelVisible(True)
            slice_data.setLabelPosition(QPieSlice.LabelOutside)
            series.append(slice_data)

        chart.addSeries(series)
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)

        chart_view.setChart(chart)
        layout.addWidget(chart_view)

        return container

    def _create_sales_tab(self) -> QWidget:
        """تب گزارش فروش با تاریخ شمسی"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        # فیلترها با تاریخ شمسی
        filter_widget = QFrame()
        filter_widget.setObjectName("filterFrame")
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setSpacing(8)
        filter_layout.setContentsMargins(8, 8, 8, 8)

        filter_layout.addWidget(QLabel("از تاریخ:"))
        self.sales_date_from = JalaliDateEdit()
        today = jdatetime.date.today()
        last_month = today - jdatetime.timedelta(days=30)
        self.sales_date_from.set_date(last_month)
        filter_layout.addWidget(self.sales_date_from)

        filter_layout.addWidget(QLabel("تا تاریخ:"))
        self.sales_date_to = JalaliDateEdit()
        self.sales_date_to.set_date(today)
        filter_layout.addWidget(self.sales_date_to)

        filter_layout.addWidget(QLabel("وضعیت:"))
        self.sales_status_combo = QComboBox()
        self.sales_status_combo.addItems(["همه", "تأیید شده", "پرداخت شده", "لغو شده"])
        filter_layout.addWidget(self.sales_status_combo)

        filter_layout.addStretch()

        self.btn_sales_filter = QPushButton("🔍 اعمال فیلتر")
        self.btn_sales_filter.setObjectName("btnPrimary")
        self.btn_sales_filter.clicked.connect(self._load_sales_report)
        filter_layout.addWidget(self.btn_sales_filter)

        layout.addWidget(filter_widget)

        # جدول فروش
        self.sales_table = QTableWidget()
        self.sales_table.setObjectName("reportTable")
        self.sales_table.setColumnCount(8)
        self.sales_table.setHorizontalHeaderLabels([
            "شماره", "تاریخ", "مشتری", "فروشنده",
            "تعداد اقلام", "جمع کل", "تخفیف", "مبلغ نهایی"
        ])
        self.sales_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sales_table.setAlternatingRowColors(True)
        layout.addWidget(self.sales_table, 1)

        # خلاصه فروش
        summary_frame = QFrame()
        summary_frame.setObjectName("summaryCard")
        summary_layout = QHBoxLayout(summary_frame)
        summary_layout.setContentsMargins(12, 8, 12, 8)

        self.sales_total_label = QLabel("💰 جمع کل فروش: ۰ ریال")
        self.sales_total_label.setStyleSheet("font-weight: bold; color: #5b3ec8; font-size: 14px;")
        summary_layout.addWidget(self.sales_total_label)

        self.sales_count_label = QLabel("📄 تعداد فاکتورها: ۰")
        self.sales_count_label.setStyleSheet("font-weight: bold; color: #4a3a6a; font-size: 14px;")
        summary_layout.addWidget(self.sales_count_label)

        summary_layout.addStretch()

        btn_print_sales = QPushButton("🖨 چاپ گزارش")
        btn_print_sales.setObjectName("btnPrimary")
        btn_print_sales.clicked.connect(lambda: self._print_report("sales"))
        summary_layout.addWidget(btn_print_sales)

        layout.addWidget(summary_frame)

        return tab

    def _create_purchases_tab(self) -> QWidget:
        """تب گزارش خرید با استفاده از جدول purchase_invoices"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        # فیلترها با تاریخ شمسی
        filter_widget = QFrame()
        filter_widget.setObjectName("filterFrame")
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setSpacing(8)
        filter_layout.setContentsMargins(8, 8, 8, 8)

        filter_layout.addWidget(QLabel("از تاریخ:"))
        self.purchases_date_from = JalaliDateEdit()
        today = jdatetime.date.today()
        last_month = today - jdatetime.timedelta(days=30)
        self.purchases_date_from.set_date(last_month)
        filter_layout.addWidget(self.purchases_date_from)

        filter_layout.addWidget(QLabel("تا تاریخ:"))
        self.purchases_date_to = JalaliDateEdit()
        self.purchases_date_to.set_date(today)
        filter_layout.addWidget(self.purchases_date_to)

        filter_layout.addWidget(QLabel("تأمین‌کننده:"))
        self.supplier_combo = QComboBox()
        self.supplier_combo.addItem("همه")
        self._load_suppliers()
        filter_layout.addWidget(self.supplier_combo)

        filter_layout.addStretch()

        self.btn_purchases_filter = QPushButton("🔍 اعمال فیلتر")
        self.btn_purchases_filter.setObjectName("btnPrimary")
        self.btn_purchases_filter.clicked.connect(self._load_purchases_report)
        filter_layout.addWidget(self.btn_purchases_filter)

        layout.addWidget(filter_widget)

        # جدول خرید
        self.purchases_table = QTableWidget()
        self.purchases_table.setObjectName("reportTable")
        self.purchases_table.setColumnCount(7)
        self.purchases_table.setHorizontalHeaderLabels([
            "شماره", "تاریخ", "تأمین‌کننده", "تعداد اقلام",
            "جمع خرید", "مالیات", "مبلغ نهایی"
        ])
        self.purchases_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.purchases_table.setAlternatingRowColors(True)
        layout.addWidget(self.purchases_table, 1)

        # خلاصه خرید
        summary_frame = QFrame()
        summary_frame.setObjectName("summaryCard")
        summary_layout = QHBoxLayout(summary_frame)
        summary_layout.setContentsMargins(12, 8, 12, 8)

        self.purchases_total_label = QLabel("💰 جمع کل خرید: ۰ ریال")
        self.purchases_total_label.setStyleSheet("font-weight: bold; color: #2e7d32; font-size: 14px;")
        summary_layout.addWidget(self.purchases_total_label)

        self.purchases_count_label = QLabel("📄 تعداد فاکتورها: ۰")
        self.purchases_count_label.setStyleSheet("font-weight: bold; color: #4a3a6a; font-size: 14px;")
        summary_layout.addWidget(self.purchases_count_label)

        summary_layout.addStretch()

        btn_print_purchases = QPushButton("🖨 چاپ گزارش")
        btn_print_purchases.setObjectName("btnPrimary")
        btn_print_purchases.clicked.connect(lambda: self._print_report("purchases"))
        summary_layout.addWidget(btn_print_purchases)

        layout.addWidget(summary_frame)

        return tab

    def _create_inventory_tab(self) -> QWidget:
        """تب گزارش انبار"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        filter_widget = QFrame()
        filter_widget.setObjectName("filterFrame")
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setSpacing(8)
        filter_layout.setContentsMargins(8, 8, 8, 8)

        filter_layout.addWidget(QLabel("دسته‌بندی:"))
        self.category_combo = QComboBox()
        self.category_combo.addItem("همه")
        self._load_categories()
        filter_layout.addWidget(self.category_combo)

        filter_layout.addWidget(QLabel("وضعیت موجودی:"))
        self.stock_status_combo = QComboBox()
        self.stock_status_combo.addItems(["همه", "کم‌تر از حد مجاز", "در حال اتمام", "موجود"])
        filter_layout.addWidget(self.stock_status_combo)

        filter_layout.addStretch()

        self.btn_inventory_filter = QPushButton("🔍 اعمال فیلتر")
        self.btn_inventory_filter.setObjectName("btnPrimary")
        self.btn_inventory_filter.clicked.connect(self._load_inventory_report)
        filter_layout.addWidget(self.btn_inventory_filter)

        layout.addWidget(filter_widget)

        self.inventory_table = QTableWidget()
        self.inventory_table.setObjectName("reportTable")
        self.inventory_table.setColumnCount(7)
        self.inventory_table.setHorizontalHeaderLabels([
            "کد", "نام کالا", "دسته‌بندی", "موجودی فعلی",
            "قیمت خرید", "قیمت فروش", "وضعیت"
        ])
        self.inventory_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.inventory_table.setAlternatingRowColors(True)
        layout.addWidget(self.inventory_table, 1)

        summary_frame = QFrame()
        summary_frame.setObjectName("summaryCard")
        summary_layout = QHBoxLayout(summary_frame)
        summary_layout.setContentsMargins(12, 8, 12, 8)

        self.inventory_total_label = QLabel("📦 تعداد کل کالاها: ۰")
        self.inventory_total_label.setStyleSheet("font-weight: bold; color: #4a3a6a; font-size: 14px;")
        summary_layout.addWidget(self.inventory_total_label)

        self.inventory_value_label = QLabel("💰 ارزش کل موجودی: ۰ ریال")
        self.inventory_value_label.setStyleSheet("font-weight: bold; color: #c62828; font-size: 14px;")
        summary_layout.addWidget(self.inventory_value_label)

        summary_layout.addStretch()

        btn_print_inventory = QPushButton("🖨 چاپ گزارش")
        btn_print_inventory.setObjectName("btnPrimary")
        btn_print_inventory.clicked.connect(lambda: self._print_report("inventory"))
        summary_layout.addWidget(btn_print_inventory)

        layout.addWidget(summary_frame)

        return tab

    def _create_products_tab(self) -> QWidget:
        """تب گزارش محصولات"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        filter_widget = QFrame()
        filter_widget.setObjectName("filterFrame")
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setSpacing(8)
        filter_layout.setContentsMargins(8, 8, 8, 8)

        filter_layout.addWidget(QLabel("جستجو:"))
        self.product_search = QLineEdit()
        self.product_search.setPlaceholderText("نام یا بارکد کالا...")
        self.product_search.setObjectName("searchInput")
        filter_layout.addWidget(self.product_search)

        filter_layout.addStretch()

        self.btn_products_search = QPushButton("🔍 جستجو")
        self.btn_products_search.setObjectName("btnPrimary")
        self.btn_products_search.clicked.connect(self._load_products_report)
        filter_layout.addWidget(self.btn_products_search)

        layout.addWidget(filter_widget)

        self.products_table = QTableWidget()
        self.products_table.setObjectName("reportTable")
        self.products_table.setColumnCount(7)
        self.products_table.setHorizontalHeaderLabels([
            "کد", "نام کالا", "بارکد", "دسته‌بندی",
            "موجودی", "قیمت خرید", "قیمت فروش"
        ])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.products_table.setAlternatingRowColors(True)
        layout.addWidget(self.products_table, 1)

        return tab

    def _create_customers_tab(self) -> QWidget:
        """تب گزارش مشتریان"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)

        filter_widget = QFrame()
        filter_widget.setObjectName("filterFrame")
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setSpacing(8)
        filter_layout.setContentsMargins(8, 8, 8, 8)

        filter_layout.addWidget(QLabel("جستجو:"))
        self.customer_search = QLineEdit()
        self.customer_search.setPlaceholderText("نام یا تلفن مشتری...")
        self.customer_search.setObjectName("searchInput")
        filter_layout.addWidget(self.customer_search)

        filter_layout.addStretch()

        self.btn_customers_search = QPushButton("🔍 جستجو")
        self.btn_customers_search.setObjectName("btnPrimary")
        self.btn_customers_search.clicked.connect(self._load_customers_report)
        filter_layout.addWidget(self.btn_customers_search)

        layout.addWidget(filter_widget)

        self.customers_table = QTableWidget()
        self.customers_table.setObjectName("reportTable")
        self.customers_table.setColumnCount(5)
        self.customers_table.setHorizontalHeaderLabels([
            "کد", "نام مشتری", "تلفن", "تعداد خرید", "مجموع خرید"
        ])
        self.customers_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.customers_table.setAlternatingRowColors(True)
        layout.addWidget(self.customers_table, 1)

        return tab

    # ========================================================================
    # توابع بارگذاری داده
    # ========================================================================

    def _load_initial_data(self):
        """بارگذاری اولیه داده‌ها"""
        self._load_sales_report()
        self._load_purchases_report()
        self._load_inventory_report()
        self._load_products_report()
        self._load_customers_report()

    def _refresh_all(self):
        """بروزرسانی همه گزارش‌ها"""
        self._load_initial_data()
        QMessageBox.information(self, "موفق", "همه گزارش‌ها بروزرسانی شدند.")

    def _get_total_sales(self) -> str:
        try:
            with _connect() as conn:
                row = conn.execute("SELECT COALESCE(SUM(total), 0) as total FROM invoices").fetchone()
                return _money(row["total"]) + " ریال"
        except Exception:
            return "۰ ریال"

    def _get_total_purchases(self) -> str:
        try:
            with _connect() as conn:
                row = conn.execute("SELECT COUNT(*) as count FROM purchase_invoices").fetchone()
                return f"{row['count']} خرید"
        except Exception:
            return "۰ خرید"

    def _get_total_products(self) -> str:
        try:
            with _connect() as conn:
                row = conn.execute("SELECT COUNT(*) as count FROM products").fetchone()
                return f"{row['count']} کالا"
        except Exception:
            return "۰ کالا"

    def _get_total_customers(self) -> str:
        try:
            with _connect() as conn:
                row = conn.execute("SELECT COUNT(*) as count FROM customers").fetchone()
                return f"{row['count']} مشتری"
        except Exception:
            return "۰ مشتری"

    def _get_monthly_sales(self) -> List[Tuple[str, float]]:
        months = []
        today = jdatetime.date.today()
        
        for i in range(11, -1, -1):
            month = today.month - i
            year = today.year
            if month <= 0:
                month += 12
                year -= 1
            try:
                jalali_date = jdatetime.date(year, month, 1)
                month_name = f"{jalali_date.strftime('%B')} {jalali_date.year}"
                months.append((month_name, 0))
            except Exception:
                months.append((f"{month:02d}/{year:02d}", 0))

        try:
            with _connect() as conn:
                rows = conn.execute("""
                    SELECT strftime('%m/%y', created_at) as month,
                           COALESCE(SUM(total), 0) as total
                    FROM invoices
                    WHERE created_at >= date('now', '-11 months')
                    GROUP BY month
                    ORDER BY month
                """).fetchall()

            for row in rows:
                try:
                    month_parts = row["month"].split('/')
                    if len(month_parts) == 2:
                        miladi_month = int(month_parts[0])
                        miladi_year = int(month_parts[1])
                        miladi_date = datetime(int(miladi_year) + 2000, miladi_month, 1)
                        jalali_date = jdatetime.datetime.fromgregorian(datetime=miladi_date)
                        month_name = f"{jalali_date.strftime('%B')} {jalali_date.year}"
                        
                        for i, (m, _) in enumerate(months):
                            if m == month_name:
                                months[i] = (m, row["total"])
                                break
                except Exception:
                    pass
        except Exception:
            pass

        return months

    def _get_products_distribution(self) -> List[Tuple[str, int]]:
        try:
            with _connect() as conn:
                rows = conn.execute("""
                    SELECT name, COALESCE(quantity, stock, 0) as qty
                    FROM products
                    WHERE COALESCE(quantity, stock, 0) > 0
                    ORDER BY qty DESC
                    LIMIT 6
                """).fetchall()
                return [(row["name"], row["qty"]) for row in rows]
        except Exception:
            return []

    def _load_suppliers(self):
        try:
            with _connect() as conn:
                rows = conn.execute("""
                    SELECT DISTINCT supplier
                    FROM purchase_invoices
                    WHERE supplier IS NOT NULL AND supplier != ''
                    ORDER BY supplier
                """).fetchall()
                for row in rows:
                    self.supplier_combo.addItem(row["supplier"])
        except Exception:
            pass

    def _load_categories(self):
        try:
            with _connect() as conn:
                rows = conn.execute("""
                    SELECT DISTINCT name
                    FROM categories
                    WHERE name IS NOT NULL AND name != ''
                    ORDER BY name
                """).fetchall()
                for row in rows:
                    self.category_combo.addItem(row["name"])
        except Exception:
            pass

    def _load_sales_report(self):
        """بارگذاری گزارش فروش با فیلتر تاریخ شمسی و وضعیت"""
        try:
            # دریافت تاریخ‌ها
            date_from = self.sales_date_from.jalali_date()
            date_to = self.sales_date_to.jalali_date()
            
            from_str = _jalali_to_gregorian_string(date_from) if date_from else "2000-01-01"
            to_str = _jalali_to_gregorian_string(date_to) if date_to else "2099-12-31"
            
            status_text = self.sales_status_combo.currentText()
            
            print(f"🔍 فیلتر فروش: از {from_str} تا {to_str} - وضعیت: {status_text}")

            with _connect() as conn:
                query = """
                    SELECT 
                        i.id,
                        i.created_at,
                        COALESCE(c.full_name, 'مشتری متفرقه') as customer,
                        COALESCE(u.full_name, 'سیستم') as seller,
                        (SELECT COUNT(*) FROM invoice_items WHERE invoice_id = i.id) as items_count,
                        i.total,
                        COALESCE(i.discount, 0) as discount,
                        i.total - COALESCE(i.discount, 0) as final_total
                    FROM invoices i
                    LEFT JOIN customers c ON c.id = i.customer_id
                    LEFT JOIN users u ON u.id = i.user_id
                    WHERE date(i.created_at) BETWEEN ? AND ?
                    ORDER BY i.created_at DESC
                    LIMIT 200
                """
                rows = conn.execute(query, (from_str, to_str)).fetchall()

                # پاک کردن جدول
                self.sales_table.setRowCount(0)
                total_sales = 0
                row_count = 0

                for r, row in enumerate(rows):
                    row_count += 1
                    
                    # تعیین وضعیت
                    status_display = "تأیید شده"
                    if row["discount"] and row["discount"] > 0:
                        status_display = "تخفیف دار"
                    if row["total"] <= 0:
                        status_display = "لغو شده"
                    
                    # فیلتر بر اساس وضعیت
                    if status_text != "همه":
                        if status_text == "تأیید شده" and status_display != "تأیید شده":
                            continue
                        if status_text == "پرداخت شده" and status_display != "تأیید شده":
                            continue
                        if status_text == "لغو شده" and status_display != "لغو شده":
                            continue
                    
                    jalali_date = _to_jalali_datetime(row["created_at"]) if row["created_at"] else ""
                    
                    values = [
                        str(row["id"]),
                        jalali_date,
                        row["customer"],
                        row["seller"],
                        str(row["items_count"]),
                        _money(row["total"] or 0),
                        _money(row["discount"] or 0),
                        _money(row["final_total"] or 0),
                    ]
                    
                    row_idx = self.sales_table.rowCount()
                    self.sales_table.insertRow(row_idx)
                    for c, val in enumerate(values):
                        item = QTableWidgetItem(val)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.sales_table.setItem(row_idx, c, item)

                    total_sales += row["total"] or 0

                self.sales_total_label.setText(f"💰 جمع کل فروش: {_money(total_sales)} ریال")
                self.sales_count_label.setText(f"📄 تعداد فاکتورها: {self.sales_table.rowCount()}")
                
                print(f"✅ گزارش فروش بارگذاری شد: {self.sales_table.rowCount()} رکورد")

        except Exception as e:
            print(f"❌ Error loading sales report: {e}")
            self.sales_table.setRowCount(0)
            self.sales_total_label.setText("💰 جمع کل فروش: ۰ ریال")
            self.sales_count_label.setText("📄 تعداد فاکتورها: ۰")
            QMessageBox.warning(self, "خطا", f"خطا در بارگذاری گزارش فروش:\n{str(e)}")

    def _load_purchases_report(self):
        """بارگذاری گزارش خرید از جدول purchase_invoices"""
        try:
            date_from = self.purchases_date_from.jalali_date()
            date_to = self.purchases_date_to.jalali_date()
            
            from_str = _jalali_to_gregorian_string(date_from) if date_from else "2000-01-01"
            to_str = _jalali_to_gregorian_string(date_to) if date_to else "2099-12-31"

            supplier = self.supplier_combo.currentText()
            supplier_filter = "" if supplier == "همه" else f"AND supplier = '{supplier}'"

            print(f"🔍 فیلتر خرید: از {from_str} تا {to_str} - تأمین‌کننده: {supplier}")

            with _connect() as conn:
                query = f"""
                    SELECT 
                        pi.id,
                        pi.date as created_at,
                        COALESCE(pi.supplier, 'نامشخص') as supplier,
                        (SELECT COUNT(*) FROM purchase_items WHERE purchase_id = pi.id) as items_count,
                        pi.total,
                        0 as tax,
                        pi.total as final_total
                    FROM purchase_invoices pi
                    WHERE date(pi.date) BETWEEN ? AND ?
                    {supplier_filter}
                    ORDER BY pi.date DESC
                    LIMIT 200
                """
                rows = conn.execute(query, (from_str, to_str)).fetchall()

                self.purchases_table.setRowCount(0)
                total_purchases = 0

                for r, row in enumerate(rows):
                    jalali_date = _to_jalali_datetime(row["created_at"]) if row["created_at"] else ""
                    
                    values = [
                        str(row["id"]),
                        jalali_date,
                        row["supplier"],
                        str(row["items_count"]),
                        _money(row["total"] or 0),
                        _money(row["tax"] or 0),
                        _money(row["final_total"] or 0),
                    ]
                    
                    row_idx = self.purchases_table.rowCount()
                    self.purchases_table.insertRow(row_idx)
                    for c, val in enumerate(values):
                        item = QTableWidgetItem(val)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.purchases_table.setItem(row_idx, c, item)

                    total_purchases += row["total"] or 0

                self.purchases_total_label.setText(f"💰 جمع کل خرید: {_money(total_purchases)} ریال")
                self.purchases_count_label.setText(f"📄 تعداد فاکتورها: {self.purchases_table.rowCount()}")
                
                print(f"✅ گزارش خرید بارگذاری شد: {self.purchases_table.rowCount()} رکورد")

        except Exception as e:
            print(f"❌ Error loading purchases report: {e}")
            self.purchases_table.setRowCount(0)
            self.purchases_total_label.setText("💰 جمع کل خرید: ۰ ریال")
            self.purchases_count_label.setText("📄 تعداد فاکتورها: ۰")
            QMessageBox.warning(self, "خطا", f"خطا در بارگذاری گزارش خرید:\n{str(e)}")

    def _load_inventory_report(self):
        """بارگذاری گزارش انبار"""
        try:
            stock_status = self.stock_status_combo.currentText()
            
            print(f"🔍 فیلتر انبار: وضعیت موجودی - {stock_status}")
            
            with _connect() as conn:
                query = """
                    SELECT 
                        p.id,
                        p.name,
                        COALESCE(c.name, 'دسته‌بندی نشده') as category,
                        COALESCE(p.quantity, p.stock, 0) as quantity,
                        COALESCE(p.unit_price, p.price, 0) as unit_price,
                        COALESCE(p.retail_price, 0) as retail_price
                    FROM products p
                    LEFT JOIN categories c ON c.id = p.category_id
                    WHERE 1=1
                """
                
                if stock_status == "کم‌تر از حد مجاز":
                    query += " AND COALESCE(p.quantity, p.stock, 0) <= 5 AND COALESCE(p.quantity, p.stock, 0) > 0"
                elif stock_status == "در حال اتمام":
                    query += " AND COALESCE(p.quantity, p.stock, 0) <= 10 AND COALESCE(p.quantity, p.stock, 0) > 5"
                elif stock_status == "موجود":
                    query += " AND COALESCE(p.quantity, p.stock, 0) > 10"
                
                query += " ORDER BY p.name"
                
                rows = conn.execute(query).fetchall()

                self.inventory_table.setRowCount(0)
                total_value = 0
                total_items = 0

                for r, row in enumerate(rows):
                    qty = row["quantity"] or 0
                    price = row["unit_price"] or 0
                    total_value += qty * price
                    total_items += qty

                    if qty <= 0:
                        status = "🔴 تمام شده"
                        status_color = "#c62828"
                    elif qty <= 5:
                        status = "🟡 کم‌تر از حد مجاز"
                        status_color = "#e65100"
                    elif qty <= 10:
                        status = "🟠 در حال اتمام"
                        status_color = "#f57c00"
                    else:
                        status = "🟢 موجود"
                        status_color = "#2e7d32"

                    values = [
                        str(row["id"]),
                        row["name"],
                        row["category"],
                        str(qty),
                        _money(price),
                        _money(row["retail_price"] or 0),
                        status,
                    ]
                    
                    row_idx = self.inventory_table.rowCount()
                    self.inventory_table.insertRow(row_idx)
                    for c, val in enumerate(values):
                        item = QTableWidgetItem(val)
                        item.setTextAlignment(Qt.AlignCenter)
                        if c == 6:
                            item.setForeground(QColor(status_color))
                        self.inventory_table.setItem(row_idx, c, item)

                self.inventory_total_label.setText(f"📦 تعداد کل کالاها: {total_items}")
                self.inventory_value_label.setText(f"💰 ارزش کل موجودی: {_money(total_value)} ریال")
                
                print(f"✅ گزارش انبار بارگذاری شد: {self.inventory_table.rowCount()} رکورد")

        except Exception as e:
            print(f"❌ Error loading inventory report: {e}")
            self.inventory_table.setRowCount(0)
            self.inventory_total_label.setText("📦 تعداد کل کالاها: ۰")
            self.inventory_value_label.setText("💰 ارزش کل موجودی: ۰ ریال")
            QMessageBox.warning(self, "خطا", f"خطا در بارگذاری گزارش انبار:\n{str(e)}")

    def _load_products_report(self):
        try:
            search = self.product_search.text().strip()
            print(f"🔍 جستجوی محصولات: {search}")
            
            with _connect() as conn:
                if search:
                    query = """
                        SELECT 
                            p.id,
                            p.name,
                            COALESCE(p.barcode, '-') as barcode,
                            COALESCE(c.name, 'دسته‌بندی نشده') as category,
                            COALESCE(p.quantity, p.stock, 0) as quantity,
                            COALESCE(p.unit_price, p.price, 0) as unit_price,
                            COALESCE(p.retail_price, 0) as retail_price
                        FROM products p
                        LEFT JOIN categories c ON c.id = p.category_id
                        WHERE p.name LIKE ? OR p.barcode LIKE ?
                        ORDER BY p.name
                        LIMIT 200
                    """
                    rows = conn.execute(query, (f"%{search}%", f"%{search}%")).fetchall()
                else:
                    query = """
                        SELECT 
                            p.id,
                            p.name,
                            COALESCE(p.barcode, '-') as barcode,
                            COALESCE(c.name, 'دسته‌بندی نشده') as category,
                            COALESCE(p.quantity, p.stock, 0) as quantity,
                            COALESCE(p.unit_price, p.price, 0) as unit_price,
                            COALESCE(p.retail_price, 0) as retail_price
                        FROM products p
                        LEFT JOIN categories c ON c.id = p.category_id
                        ORDER BY p.name
                        LIMIT 200
                    """
                    rows = conn.execute(query).fetchall()

                self.products_table.setRowCount(0)

                for r, row in enumerate(rows):
                    values = [
                        str(row["id"]),
                        row["name"],
                        row["barcode"],
                        row["category"],
                        str(row["quantity"] or 0),
                        _money(row["unit_price"] or 0),
                        _money(row["retail_price"] or 0),
                    ]
                    
                    row_idx = self.products_table.rowCount()
                    self.products_table.insertRow(row_idx)
                    for c, val in enumerate(values):
                        item = QTableWidgetItem(val)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.products_table.setItem(row_idx, c, item)

                print(f"✅ گزارش محصولات بارگذاری شد: {self.products_table.rowCount()} رکورد")

        except Exception as e:
            print(f"❌ Error loading products report: {e}")
            self.products_table.setRowCount(0)
            QMessageBox.warning(self, "خطا", f"خطا در جستجوی محصولات:\n{str(e)}")

    def _load_customers_report(self):
        try:
            search = self.customer_search.text().strip()
            print(f"🔍 جستجوی مشتریان: {search}")
            
            with _connect() as conn:
                if search:
                    query = """
                        SELECT 
                            c.id,
                            c.full_name as name,
                            COALESCE(c.phone, '-') as phone,
                            COUNT(DISTINCT i.id) as purchase_count,
                            COALESCE(SUM(i.total), 0) as total_purchases
                        FROM customers c
                        LEFT JOIN invoices i ON i.customer_id = c.id
                        WHERE c.full_name LIKE ? OR c.phone LIKE ?
                        GROUP BY c.id
                        ORDER BY total_purchases DESC
                        LIMIT 200
                    """
                    rows = conn.execute(query, (f"%{search}%", f"%{search}%")).fetchall()
                else:
                    query = """
                        SELECT 
                            c.id,
                            c.full_name as name,
                            COALESCE(c.phone, '-') as phone,
                            COUNT(DISTINCT i.id) as purchase_count,
                            COALESCE(SUM(i.total), 0) as total_purchases
                        FROM customers c
                        LEFT JOIN invoices i ON i.customer_id = c.id
                        GROUP BY c.id
                        ORDER BY total_purchases DESC
                        LIMIT 200
                    """
                    rows = conn.execute(query).fetchall()

                self.customers_table.setRowCount(0)

                for r, row in enumerate(rows):
                    values = [
                        str(row["id"]),
                        row["name"],
                        row["phone"],
                        str(row["purchase_count"] or 0),
                        _money(row["total_purchases"] or 0),
                    ]
                    
                    row_idx = self.customers_table.rowCount()
                    self.customers_table.insertRow(row_idx)
                    for c, val in enumerate(values):
                        item = QTableWidgetItem(val)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.customers_table.setItem(row_idx, c, item)

                print(f"✅ گزارش مشتریان بارگذاری شد: {self.customers_table.rowCount()} رکورد")

        except Exception as e:
            print(f"❌ Error loading customers report: {e}")
            self.customers_table.setRowCount(0)
            QMessageBox.warning(self, "خطا", f"خطا در جستجوی مشتریان:\n{str(e)}")

    def _export_excel(self):
        """خروجی Excel از گزارش جاری"""
        current_tab = self.tabs.currentIndex()
        tab_names = ["dashboard", "sales", "purchases", "inventory", "products", "customers"]
        tab_name = tab_names[current_tab] if current_tab < len(tab_names) else "report"

        jalali_now = jdatetime.datetime.now()
        date_str = jalali_now.strftime("%Y%m%d")

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره گزارش Excel",
            f"report_{tab_name}_{date_str}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if file_path:
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "گزارش"
                
                table = self._get_current_table()
                if table:
                    headers = table.horizontalHeader()
                    for col in range(headers.count()):
                        ws.cell(row=1, column=col+1, value=headers.model().headerData(col, Qt.Horizontal))
                    
                    for row in range(table.rowCount()):
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            if item:
                                ws.cell(row=row+2, column=col+1, value=item.text())
                
                wb.save(file_path)
                QMessageBox.information(self, "موفق", f"گزارش در مسیر زیر ذخیره شد:\n{file_path}")
                _open_file(file_path)
                
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در ذخیره Excel:\n{str(e)}")

    def _export_pdf(self):
        """خروجی PDF از گزارش جاری"""
        current_tab = self.tabs.currentIndex()
        tab_names = ["dashboard", "sales", "purchases", "inventory", "products", "customers"]
        tab_name = tab_names[current_tab] if current_tab < len(tab_names) else "report"

        jalali_now = jdatetime.datetime.now()
        date_str = jalali_now.strftime("%Y%m%d")

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره گزارش PDF",
            f"report_{tab_name}_{date_str}.pdf",
            "PDF Files (*.pdf)"
        )

        if file_path:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                
                doc = SimpleDocTemplate(file_path, pagesize=A4)
                story = []
                styles = getSampleStyleSheet()
                
                title_style = ParagraphStyle(
                    'TitleStyle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    alignment=1,
                    spaceAfter=12
                )
                story.append(Paragraph(f"گزارش {tab_name}", title_style))
                story.append(Spacer(1, 0.5*cm))
                
                date_style = ParagraphStyle(
                    'DateStyle',
                    parent=styles['Normal'],
                    fontSize=10,
                    alignment=2
                )
                story.append(Paragraph(f"تاریخ: {jalali_now.strftime('%Y/%m/%d %H:%M')}", date_style))
                story.append(Spacer(1, 0.5*cm))
                
                table = self._get_current_table()
                if table:
                    data = []
                    headers = table.horizontalHeader()
                    header_row = []
                    for col in range(headers.count()):
                        header_row.append(headers.model().headerData(col, Qt.Horizontal))
                    data.append(header_row)
                    
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data.append(row_data)
                    
                    tbl = Table(data, repeatRows=1)
                    tbl.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8e0f5')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4a3a6a')),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    story.append(tbl)
                
                doc.build(story)
                QMessageBox.information(self, "موفق", f"گزارش در مسیر زیر ذخیره شد:\n{file_path}")
                _open_file(file_path)
                
            except ImportError:
                QMessageBox.warning(self, "خطا", "لطفاً کتابخانه reportlab را نصب کنید:\npip install reportlab")
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در ذخیره PDF:\n{str(e)}")

    def _get_current_table(self) -> Optional[QTableWidget]:
        """دریافت جدول تب جاری"""
        current_tab = self.tabs.currentIndex()
        if current_tab == 1:  # فروش
            return self.sales_table
        elif current_tab == 2:  # خرید
            return self.purchases_table
        elif current_tab == 3:  # انبار
            return self.inventory_table
        elif current_tab == 4:  # محصولات
            return self.products_table
        elif current_tab == 5:  # مشتریان
            return self.customers_table
        return None

    def _print_report(self, report_type: str):
        """چاپ گزارش"""
        QMessageBox.information(self, "چاپ", f"در حال چاپ گزارش {report_type}...")

    # ========================================================================
    # استایل‌ها
    # ========================================================================

    def _apply_styles(self):
        self.setStyleSheet("""
            #pageTitle {
                color: #5B4A8A;
                font-size: 22px;
                font-weight: bold;
                padding: 8px 0;
            }

            #btnPrimary {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #8a6ed6, stop:1 #6a4c9c);
                color: #ffffff;
                border: none;
                border-radius: 12px;
                padding: 8px 20px;
                font-weight: bold;
                font-size: 13px;
            }
            #btnPrimary:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #9a7ee6, stop:1 #7a5cac);
            }

            #reportTabs::pane {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 16px;
                padding: 12px;
            }
            #reportTabs QTabBar::tab {
                padding: 10px 24px;
                border-radius: 12px 12px 0 0;
                font-size: 13px;
                font-weight: 600;
                color: #6a5a8a;
                background: transparent;
                margin-right: 4px;
            }
            #reportTabs QTabBar::tab:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #d5c8ed, stop:1 #e8e0f5);
                color: #3a2a5a;
            }
            #reportTabs QTabBar::tab:hover:!selected {
                background: rgba(213, 200, 237, 0.3);
                border-radius: 12px 12px 0 0;
            }

            #reportTable {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 14px;
                gridline-color: #f0ecf8;
            }
            #reportTable::item {
                padding: 8px;
                border-radius: 8px;
                font-size: 13px;
            }
            #reportTable::item:selected {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #d5c8ed, stop:1 #e8e0f5);
                color: #3a2a5a;
                border-radius: 8px;
            }
            #reportTable QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #e8e0f5, stop:1 #d5c8ed);
                color: #4a3a6a;
                font-weight: bold;
                border: none;
                border-right: 1px solid #f0ecf8;
                padding: 10px 6px;
                font-size: 12px;
            }
            #reportTable QHeaderView::section:last {
                border-right: none;
            }

            #summaryCard {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 14px;
                padding: 12px 18px;
            }

            #filterFrame {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 14px;
                padding: 4px;
            }

            #searchInput {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 14px;
                padding: 8px 14px;
                font-size: 13px;
                color: #4a3a6a;
                min-width: 200px;
            }
            #searchInput:focus {
                border-color: #a48ad6;
            }
            #searchInput::placeholder {
                color: #b0a8c8;
                font-style: italic;
            }

            QComboBox {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 12px;
                padding: 6px 12px;
                font-size: 13px;
                color: #4a3a6a;
                min-height: 32px;
                min-width: 120px;
            }
            QComboBox:focus {
                border-color: #a48ad6;
            }
            QComboBox::drop-down {
                border: none;
                border-radius: 8px;
            }

            /* استایل ویجت تاریخ شمسی */
            #jalaliDateLabel {
                background: #faf8ff;
                border: 2px solid #e8e0f5;
                border-radius: 12px;
                padding: 6px 12px;
                font-size: 13px;
                color: #4a3a6a;
                min-height: 32px;
                min-width: 120px;
                text-align: center;
            }
            #jalaliDateLabel:hover {
                border-color: #a48ad6;
            }

            #clearDateBtn {
                background: transparent;
                border: none;
                border-radius: 8px;
                color: #b0a8c8;
                font-size: 12px;
            }
            #clearDateBtn:hover {
                background: #f0ecf8;
                color: #c62828;
            }

            #calendarFrame {
                background: #ffffff;
                border: 2px solid #e8e0f5;
                border-radius: 14px;
                padding: 10px;
                box-shadow: 0 8px 30px rgba(0, 0, 0, 0.15);
            }

            #calendarMonthYear {
                font-size: 14px;
                font-weight: bold;
                color: #4a3a6a;
            }

            #calendarNavBtn {
                background: transparent;
                border: none;
                border-radius: 8px;
                padding: 4px 10px;
                font-size: 14px;
                color: #6a5a8a;
            }
            #calendarNavBtn:hover {
                background: #f0ecf8;
            }

            #weekdayLabel {
                font-size: 11px;
                font-weight: bold;
                color: #8a7aaa;
            }

            #dayBtn {
                background: transparent;
                border: none;
                border-radius: 8px;
                font-size: 12px;
                color: #4a3a6a;
                padding: 0px;
                min-height: 28px;
            }
            #dayBtn:hover {
                background: #e8e0f5;
            }
            #dayBtn[selected="true"] {
                background: #8a6ed6;
                color: #ffffff;
                font-weight: bold;
            }
            #dayBtn:disabled {
                color: #d5cbee;
            }

            #todayBtn {
                background: #f0ecf8;
                border: none;
                border-radius: 8px;
                padding: 6px;
                font-size: 12px;
                font-weight: bold;
                color: #4a3a6a;
            }
            #todayBtn:hover {
                background: #e8e0f5;
            }

            QLabel {
                color: #4a3a6a;
                font-size: 13px;
            }

            QScrollBar:vertical {
                background: #f5f0fa;
                border-radius: 10px;
                width: 10px;
                margin: 4px;
            }
            QScrollBar::handle:vertical {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #d5c8ed, stop:1 #c0b0e0);
                border-radius: 8px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #c0b0e0, stop:1 #a48ad6);
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)