# -*- coding: utf-8 -*-
"""صفحه حسابداری - مدیریت اسناد، ترازنامه، دفتر کل، ثبت چک‌ها و تقویم سه‌بعدی با وضوح بالا"""

import os
from PySide6.QtCore import Qt, QDate, QPoint, Signal, QUrl
from PySide6.QtGui import QFont, QCursor, QTextDocument, QDesktopServices
from PySide6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame, QMessageBox,
    QDialog, QFormLayout, QTextEdit, QDoubleSpinBox, QDialogButtonBox,
    QTabWidget, QComboBox, QToolButton, QGridLayout, QFileDialog
)

try:
    import jdatetime
    JALALI_AVAILABLE = True
except ImportError:
    JALALI_AVAILABLE = False

try:
    from persiantools.jalali import JalaliDate
    PERSIAN_TOOLS_AVAILABLE = True
except ImportError:
    PERSIAN_TOOLS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.database.connection import get_connection


def to_persian_digits(num_str: str) -> str:
    """تبدیل اعداد انگلیسی به فارسی برای خوانایی و زیبایی بیشتر"""
    persian_digits = {'0': '۰', '1': '۱', '2': '۲', '3': '۳', '4': '۴',
                      '5': '۵', '6': '۶', '7': '۷', '8': '۸', '9': '۹'}
    return "".join(persian_digits.get(c, c) for c in str(num_str))


def is_official_holiday(year: int, month: int, day: int) -> bool:
    """بررسی تعطیلی رسمی"""
    if PERSIAN_TOOLS_AVAILABLE:
        try:
            return JalaliDate(year, month, day).is_holiday()
        except:
            return False
    else:
        OFFICIAL_HOLIDAYS = [
            "01-01", "01-02", "01-03", "01-04", "01-12", "01-13",
            "02-11", "02-12", "02-13", "02-14",
            "03-14", "03-15",
            "06-25", "06-26",
            "11-22", "11-23",
            "09-08", "09-09",
            "12-29",
        ]
        return f"{month:02d}-{day:02d}" in OFFICIAL_HOLIDAYS


def is_holiday(j_date) -> bool:
    """بررسی تعطیل بودن روز (جمعه یا تعطیل رسمی)"""
    if j_date.weekday() == 6:
        return True
    return is_official_holiday(j_date.year, j_date.month, j_date.day)


class ShamsiCalendarPopup(QWidget):
    """تقویم شمسی سه‌بعدی و مدرن با اعداد کاملاً خوانا و واضح"""
    date_selected = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent, Qt.Popup)
        self.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint | Qt.NoDropShadowWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setFixedSize(340, 390)

        self.setStyleSheet("""
            QFrame#calContainer {
                background-color: #ffffff;
                border: 2px solid #7c3aed;
                border-radius: 16px;
            }
            QLabel#monthTitle {
                color: #2e1065;
                font-family: 'Segoe UI', 'Tahoma', 'Vazirmatn', sans-serif;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton#navBtn {
                background-color: #7c3aed;
                color: #ffffff !important;
                border: 1px solid #5b21b6;
                border-radius: 15px;
                font-family: 'Segoe UI', 'Tahoma';
                font-size: 15px;
                font-weight: bold;
                padding: 0px !important;
                margin: 0px;
                text-align: center;
            }
            QPushButton#navBtn:hover {
                background-color: #6d28d9;
            }
            QPushButton#navBtn:pressed {
                background-color: #4c1d95;
            }
            QLabel#weekDayHeader {
                color: #6b7280;
                font-family: 'Segoe UI', 'Tahoma', 'Vazirmatn';
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton#dayBtn {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffffff, stop:1 #f3f0fb);
                border: 1.5px solid #dcd3f5;
                border-radius: 8px;
                font-family: 'Segoe UI', 'Tahoma', 'Vazirmatn', sans-serif;
                font-size: 13px;
                font-weight: bold;
                color: #1e1b4b;
                padding: 0px !important;
                margin: 0px;
            }
            QPushButton#dayBtn:hover {
                background: #ddd6fe;
                border: 1.5px solid #7c3aed;
                color: #4c1d95;
            }
            QPushButton#dayBtn:pressed {
                background: #c4b5fd;
                padding-top: 2px !important;
            }
            QPushButton#dayBtn[today="true"] {
                background: #7c3aed !important;
                color: #ffffff !important;
                border: 1.5px solid #5b21b6 !important;
                font-weight: bold !important;
            }
            QPushButton#dayBtn[holiday="true"] {
                background: #fee2e2 !important;
                color: #b91c1c !important;
                border: 1.5px solid #fca5a5 !important;
                font-weight: bold !important;
            }
            QPushButton#dayBtn[holiday="true"]:hover {
                background: #fecaca !important;
            }
            QPushButton#todayBtn {
                background-color: #f5f3ff;
                color: #6d28d9;
                border: 1.5px solid #c4b5fd;
                border-radius: 10px;
                font-family: 'Segoe UI', 'Tahoma', 'Vazirmatn';
                font-weight: bold;
                font-size: 12px;
                padding: 6px;
            }
            QPushButton#todayBtn:hover {
                background-color: #ede9fe;
                border-color: #7c3aed;
            }
        """)

        today = jdatetime.date.today() if JALALI_AVAILABLE else None
        self.today_date = today
        self.current_year = today.year if today else 1403
        self.current_month = today.month if today else 1

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(6, 6, 6, 6)

        self.container = QFrame()
        self.container.setObjectName("calContainer")
        self.card_layout = QVBoxLayout(self.container)
        self.card_layout.setContentsMargins(12, 12, 12, 12)
        self.card_layout.setSpacing(8)

        # نوار بالای تقویم
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(8)

        self.prev_btn = QPushButton("◀")
        self.prev_btn.setObjectName("navBtn")
        self.prev_btn.setFixedSize(32, 32)
        self.prev_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.prev_btn.clicked.connect(self._prev_month)
        header_layout.addWidget(self.prev_btn)

        self.month_label = QLabel()
        self.month_label.setObjectName("monthTitle")
        self.month_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(self.month_label, 1)

        self.next_btn = QPushButton("▶")
        self.next_btn.setObjectName("navBtn")
        self.next_btn.setFixedSize(32, 32)
        self.next_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.next_btn.clicked.connect(self._next_month)
        header_layout.addWidget(self.next_btn)

        self.card_layout.addLayout(header_layout)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background-color: #e5e7eb; max-height: 1px;")
        self.card_layout.addWidget(sep)

        # روزهای هفته
        week_layout = QHBoxLayout()
        week_layout.setSpacing(2)
        for day in ["ش", "ی", "د", "س", "چ", "پ", "ج"]:
            lbl = QLabel(day)
            lbl.setObjectName("weekDayHeader")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setFixedHeight(22)
            week_layout.addWidget(lbl)
        self.card_layout.addLayout(week_layout)

        # شبکه روزها
        self.days_grid = QGridLayout()
        self.days_grid.setSpacing(4)
        self.card_layout.addLayout(self.days_grid)

        # دکمه امروز
        self.today_action_btn = QPushButton("📌 امروز")
        self.today_action_btn.setObjectName("todayBtn")
        self.today_action_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.today_action_btn.clicked.connect(self._go_today)
        self.card_layout.addWidget(self.today_action_btn)

        root_layout.addWidget(self.container)
        self._render_month()

    def _render_month(self):
        while self.days_grid.count():
            item = self.days_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        month_names = [
            "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
            "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
        ]
        year_str = to_persian_digits(str(self.current_year))
        month_str = month_names[self.current_month - 1]
        self.month_label.setText(f"{month_str}  {year_str}")

        if not JALALI_AVAILABLE:
            return

        if self.current_month <= 6:
            days_in_month = 31
        elif self.current_month <= 11:
            days_in_month = 30
        else:
            days_in_month = 30 if jdatetime.date(self.current_year, 12, 1).isleap() else 29

        first_day = jdatetime.date(self.current_year, self.current_month, 1)
        start_col = first_day.weekday()

        for day in range(1, days_in_month + 1):
            j_date = jdatetime.date(self.current_year, self.current_month, day)

            btn = QPushButton(to_persian_digits(str(day)))
            btn.setObjectName("dayBtn")
            btn.setFixedSize(38, 34)
            btn.setCursor(QCursor(Qt.PointingHandCursor))

            if self.today_date and j_date == self.today_date:
                btn.setProperty("today", "true")
            elif is_holiday(j_date):
                btn.setProperty("holiday", "true")

            btn.clicked.connect(lambda _, d=day: self._select_date(d))
            self.days_grid.addWidget(btn, (day - 1 + start_col) // 7, (day - 1 + start_col) % 7)

    def _select_date(self, day: int):
        if JALALI_AVAILABLE:
            selected = jdatetime.date(self.current_year, self.current_month, day)
            self.date_selected.emit(selected.strftime("%Y/%m/%d"))
        self.close()

    def _go_today(self):
        if self.today_date:
            self.current_year = self.today_date.year
            self.current_month = self.today_date.month
            self._render_month()

    def _prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self._render_month()

    def _next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self._render_month()


class ShamsiDateWidget(QWidget):
    """ویجت انتخاب تاریخ شمسی"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self.line_edit = QLineEdit()
        self.line_edit.setReadOnly(True)
        self.line_edit.setPlaceholderText("انتخاب تاریخ...")
        self.line_edit.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: 1.5px solid #dcd3f5;
                border-radius: 8px;
                padding: 6px 10px;
                font-family: 'Segoe UI', 'Tahoma', sans-serif;
                font-weight: bold;
                font-size: 13px;
                color: #2e1065;
            }
            QLineEdit:focus {
                border-color: #7c3aed;
            }
        """)
        self.line_edit.setText(
            jdatetime.date.today().strftime("%Y/%m/%d") if JALALI_AVAILABLE else QDate.currentDate().toString("yyyy/MM/dd")
        )
        layout.addWidget(self.line_edit, 1)

        self.cal_btn = QToolButton()
        self.cal_btn.setText("📅")
        self.cal_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.cal_btn.setFixedSize(36, 34)
        self.cal_btn.setStyleSheet("""
            QToolButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ede9fe, stop:1 #ddd6fe);
                border: 1.5px solid #c4b5fd;
                border-radius: 8px;
                font-size: 15px;
            }
            QToolButton:hover {
                background-color: #c4b5fd;
            }
            QToolButton:pressed {
                padding-top: 2px;
            }
        """)
        self.cal_btn.clicked.connect(self._show_calendar)
        layout.addWidget(self.cal_btn)

        self._calendar = ShamsiCalendarPopup(self)
        self._calendar.date_selected.connect(self._on_date_selected)

    def _show_calendar(self):
        pos = self.cal_btn.mapToGlobal(QPoint(0, self.cal_btn.height() + 5))
        self._calendar.move(pos.x() - 290, pos.y())
        self._calendar.show()

    def _on_date_selected(self, date_str: str):
        self.line_edit.setText(date_str)

    def get_jalali_str(self) -> str:
        return self.line_edit.text()

    def set_jalali_str(self, value: str):
        self.line_edit.setText(value)


class NewDocDialog(QDialog):
    """فرم ثبت یا ویرایش سند حسابداری"""
    def __init__(self, parent=None, doc_data=None):
        super().__init__(parent)
        self.doc_data = doc_data
        self.setWindowTitle("ویرایش سند حسابداری" if doc_data else "ثبت سند حسابداری جدید")
        self.setMinimumWidth(460)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout(self)

        form = QFormLayout()
        form.setSpacing(12)

        self.doc_number = QLineEdit()
        self.doc_number.setPlaceholderText("مثلاً: 1001")
        if doc_data:
            self.doc_number.setText(str(doc_data.get("doc_number", "")))
        form.addRow("شماره سند:", self.doc_number)

        self.date = ShamsiDateWidget()
        if doc_data and doc_data.get("date"):
            self.date.set_jalali_str(str(doc_data["date"]))
        form.addRow("تاریخ:", self.date)

        self.description = QTextEdit()
        self.description.setPlaceholderText("شرح سند...")
        if doc_data:
            self.description.setText(str(doc_data.get("description", "")))
        self.description.setFixedHeight(80)
        form.addRow("شرح:", self.description)

        self.debit = QDoubleSpinBox()
        self.debit.setRange(0, 999999999999)
        self.debit.setDecimals(0)
        self.debit.setSuffix(" ریال")
        if doc_data:
            self.debit.setValue(float(doc_data.get("debit", 0)))
        form.addRow("بدهکار:", self.debit)

        self.credit = QDoubleSpinBox()
        self.credit.setRange(0, 999999999999)
        self.credit.setDecimals(0)
        self.credit.setSuffix(" ریال")
        if doc_data:
            self.credit.setValue(float(doc_data.get("credit", 0)))
        form.addRow("بستانکار:", self.credit)

        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Save).setText("ذخیره سند")
        buttons.button(QDialogButtonBox.Cancel).setText("انصراف")
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)

        layout.addWidget(buttons)

    def _validate(self):
        if not self.doc_number.text().strip():
            QMessageBox.warning(self, "خطا", "شماره سند الزامی است.")
            return
        if self.debit.value() == 0 and self.credit.value() == 0:
            QMessageBox.warning(self, "خطا", "مبلغ بدهکار یا بستانکار باید بزرگتر از صفر باشد.")
            return
        self.accept()

    def get_data(self) -> dict:
        return {
            "doc_number": self.doc_number.text().strip(),
            "date": self.date.get_jalali_str(),
            "description": self.description.toPlainText().strip(),
            "debit": self.debit.value(),
            "credit": self.credit.value(),
        }


class CheckDialog(QDialog):
    """فرم ثبت یا ویرایش چک"""
    def __init__(self, parent=None, check_data=None):
        super().__init__(parent)
        self.check_data = check_data
        self.setWindowTitle("ویرایش چک" if check_data else "ثبت چک جدید")
        self.setMinimumWidth(500)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout(self)

        form = QFormLayout()
        form.setSpacing(10)

        # --- اطلاعات چک ---
        self.check_number = QLineEdit()
        self.check_number.setPlaceholderText("شماره صیادی / چک...")
        if check_data:
            self.check_number.setText(str(check_data.get("check_number", "")))
        form.addRow("شماره چک:", self.check_number)

        self.amount = QDoubleSpinBox()
        self.amount.setRange(0, 999999999999)
        self.amount.setDecimals(0)
        self.amount.setSuffix(" ریال")
        if check_data:
            self.amount.setValue(float(check_data.get("amount", 0)))
        form.addRow("مبلغ چک:", self.amount)

        self.due_date = ShamsiDateWidget()
        if check_data and check_data.get("due_date"):
            self.due_date.set_jalali_str(str(check_data["due_date"]))
        form.addRow("تاریخ سررسید:", self.due_date)

        self.bank = QLineEdit()
        self.bank.setPlaceholderText("نام بانک...")
        if check_data:
            self.bank.setText(str(check_data.get("bank", "")))
        form.addRow("بانک:", self.bank)

        self.status = QComboBox()
        self.status.addItem("دریافتی", "دریافتی")
        self.status.addItem("پرداختی", "پرداختی")
        if check_data and check_data.get("status"):
            self.status.setCurrentText(str(check_data["status"]))
        form.addRow("نوع چک:", self.status)

        self.description = QTextEdit()
        self.description.setPlaceholderText("توضیحات چک...")
        if check_data:
            self.description.setText(str(check_data.get("description", "")))
        self.description.setFixedHeight(60)
        form.addRow("توضیحات:", self.description)

        # --- اطلاعات صادرکننده چک ---
        issuer_title = QLabel("👤 اطلاعات صادرکننده چک")
        issuer_title.setStyleSheet("font-weight: bold; color: #4d3a78; margin-top: 10px;")
        form.addRow(issuer_title)

        self.issuer_name = QLineEdit()
        self.issuer_name.setPlaceholderText("نام و نام خانوادگی...")
        if check_data:
            self.issuer_name.setText(str(check_data.get("issuer_name", "")))
        form.addRow("نام و نام خانوادگی:", self.issuer_name)

        self.issuer_national_id = QLineEdit()
        self.issuer_national_id.setPlaceholderText("شماره ملی...")
        if check_data:
            self.issuer_national_id.setText(str(check_data.get("issuer_national_id", "")))
        form.addRow("شماره ملی:", self.issuer_national_id)

        self.issuer_phone = QLineEdit()
        self.issuer_phone.setPlaceholderText("شماره موبایل...")
        if check_data:
            self.issuer_phone.setText(str(check_data.get("issuer_phone", "")))
        form.addRow("شماره موبایل:", self.issuer_phone)

        # --- اطلاعات دریافت‌کننده چک ---
        receiver_title = QLabel("👥 اطلاعات دریافت‌کننده چک")
        receiver_title.setStyleSheet("font-weight: bold; color: #4d3a78; margin-top: 10px;")
        form.addRow(receiver_title)

        self.receiver_name = QLineEdit()
        self.receiver_name.setPlaceholderText("نام و نام خانوادگی...")
        if check_data:
            self.receiver_name.setText(str(check_data.get("receiver_name", "")))
        form.addRow("نام و نام خانوادگی:", self.receiver_name)

        self.receiver_national_id = QLineEdit()
        self.receiver_national_id.setPlaceholderText("شماره ملی...")
        if check_data:
            self.receiver_national_id.setText(str(check_data.get("receiver_national_id", "")))
        form.addRow("شماره ملی:", self.receiver_national_id)

        self.receiver_phone = QLineEdit()
        self.receiver_phone.setPlaceholderText("شماره موبایل...")
        if check_data:
            self.receiver_phone.setText(str(check_data.get("receiver_phone", "")))
        form.addRow("شماره موبایل:", self.receiver_phone)

        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Save).setText("ذخیره چک")
        buttons.button(QDialogButtonBox.Cancel).setText("انصراف")
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)

        layout.addWidget(buttons)

    def _validate(self):
        if not self.check_number.text().strip():
            QMessageBox.warning(self, "خطا", "شماره چک الزامی است.")
            return
        if self.amount.value() == 0:
            QMessageBox.warning(self, "خطا", "مبلغ چک الزامی است.")
            return
        self.accept()

    def get_data(self) -> dict:
        return {
            "check_number": self.check_number.text().strip(),
            "amount": self.amount.value(),
            "due_date": self.due_date.get_jalali_str(),
            "bank": self.bank.text().strip(),
            "status": self.status.currentData(),
            "description": self.description.toPlainText().strip(),
            "issuer_name": self.issuer_name.text().strip(),
            "issuer_national_id": self.issuer_national_id.text().strip(),
            "issuer_phone": self.issuer_phone.text().strip(),
            "receiver_name": self.receiver_name.text().strip(),
            "receiver_national_id": self.receiver_national_id.text().strip(),
            "receiver_phone": self.receiver_phone.text().strip(),
        }


class ChecksTab(QWidget):
    """تب مدیریت چک‌ها"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self._load_checks()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # نوار جستجو و دکمه‌ها
        toolbar = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("جستجو بر اساس شماره چک، بانک یا نام...")
        self.search_edit.setStyleSheet("padding: 7px; border-radius: 8px; border: 1px solid #cfc8ea;")
        self.search_edit.textChanged.connect(self._search_checks)
        toolbar.addWidget(self.search_edit, 1)

        self.add_btn = QPushButton("➕ ثبت چک جدید")
        self.add_btn.setStyleSheet(
            "QPushButton{background:#7c3aed;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#6d28d9;}"
        )
        self.add_btn.clicked.connect(self._add_check)
        toolbar.addWidget(self.add_btn)

        # دکمه پیش‌نمایش
        self.preview_btn = QPushButton("👁 پیش‌نمایش")
        self.preview_btn.setStyleSheet(
            "QPushButton{background:#0ea5e9;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#0284c7;}"
        )
        self.preview_btn.clicked.connect(self._preview_checks)
        toolbar.addWidget(self.preview_btn)

        # دکمه چاپ
        self.print_btn = QPushButton("🖨 چاپ")
        self.print_btn.setStyleSheet(
            "QPushButton{background:#16a34a;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#15803d;}"
        )
        self.print_btn.clicked.connect(self._print_checks)
        toolbar.addWidget(self.print_btn)

        # دکمه ذخیره PDF
        self.pdf_btn = QPushButton("📄 خروجی PDF")
        self.pdf_btn.setStyleSheet(
            "QPushButton{background:#dc2626;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#b91c1c;}"
        )
        self.pdf_btn.clicked.connect(self._export_checks_pdf)
        toolbar.addWidget(self.pdf_btn)

        # دکمه ذخیره Excel
        self.excel_btn = QPushButton("📊 خروجی Excel")
        self.excel_btn.setStyleSheet(
            "QPushButton{background:#f59e0b;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#d97706;}"
        )
        self.excel_btn.clicked.connect(self._export_checks_excel)
        toolbar.addWidget(self.excel_btn)

        layout.addLayout(toolbar)

        # جدول چک‌ها (شامل ستون عملیات)
        self.table = QTableWidget()
        self.table.setColumnCount(14)
        self.table.setHorizontalHeaderLabels([
            "شماره چک", "مبلغ", "تاریخ سررسید", "بانک", "نوع", "توضیحات",
            "صادرکننده", "شماره ملی (صادرکننده)", "موبایل (صادرکننده)",
            "دریافت‌کننده", "شماره ملی (دریافت‌کننده)", "موبایل (دریافت‌کننده)", "", ""
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.table)

    def _load_checks(self):
        try:
            with get_connection() as conn:
                rows = conn.execute("SELECT * FROM checks ORDER BY id DESC").fetchall()
                self._populate_table(rows)
        except Exception:
            self.table.setRowCount(0)

    def _populate_table(self, rows):
        """پر کردن جدول با داده‌ها"""
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            values = [
                str(row["check_number"]),
                f"{int(row['amount']):,} ریال",
                str(row["due_date"]),
                str(row["bank"]),
                str(row["status"]),
                str(row["description"] or "—"),
                str(row["issuer_name"] or "—"),
                str(row["issuer_national_id"] or "—"),
                str(row["issuer_phone"] or "—"),
                str(row["receiver_name"] or "—"),
                str(row["receiver_national_id"] or "—"),
                str(row["receiver_phone"] or "—")
            ]
            for c, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(r, c, item)

            # دکمه ویرایش
            edit_btn = QPushButton("✏️ ویرایش")
            edit_btn.setStyleSheet(
                "QPushButton{background:#0ea5e9;color:white;border-radius:6px;padding:4px 10px;font-weight:bold;border:none;}"
                "QPushButton:hover{background:#0284c7;}"
            )
            check_data = dict(row)
            edit_btn.clicked.connect(lambda _, d=check_data: self._edit_check(d))
            self.table.setCellWidget(r, 12, edit_btn)

            # دکمه حذف
            delete_btn = QPushButton("🗑️ حذف")
            delete_btn.setStyleSheet(
                "QPushButton{background:#ef4444;color:white;border-radius:6px;padding:4px 10px;font-weight:bold;border:none;}"
                "QPushButton:hover{background:#dc2626;}"
            )
            check_id = row["id"]
            delete_btn.clicked.connect(lambda _, did=check_id: self._delete_check(did))
            self.table.setCellWidget(r, 13, delete_btn)

    def _search_checks(self, text):
        """جستجو در چک‌ها"""
        term = f"%{text.strip()}%"
        try:
            with get_connection() as conn:
                rows = conn.execute(
                    """SELECT * FROM checks WHERE check_number LIKE ? OR bank LIKE ? 
                       OR issuer_name LIKE ? OR receiver_name LIKE ? ORDER BY id DESC""",
                    (term, term, term, term)
                ).fetchall()
                self._populate_table(rows)
        except Exception:
            pass

    def _add_check(self):
        dlg = CheckDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return

        data = dlg.get_data()
        try:
            with get_connection() as conn:
                conn.execute(
                    """INSERT INTO checks 
                    (check_number, amount, due_date, bank, status, description, 
                     issuer_name, issuer_national_id, issuer_phone, 
                     receiver_name, receiver_national_id, receiver_phone) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        data["check_number"], data["amount"], data["due_date"],
                        data["bank"], data["status"], data["description"],
                        data["issuer_name"], data["issuer_national_id"], data["issuer_phone"],
                        data["receiver_name"], data["receiver_national_id"], data["receiver_phone"]
                    )
                )
            QMessageBox.information(self, "موفق", "چک با موفقیت ثبت شد.")
            self._load_checks()
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ثبت چک:\n{str(e)}")

    def _edit_check(self, check_data):
        """ویرایش چک"""
        dlg = CheckDialog(self, check_data=check_data)
        if dlg.exec() != QDialog.Accepted:
            return

        data = dlg.get_data()
        check_id = check_data["id"]
        try:
            with get_connection() as conn:
                conn.execute(
                    """UPDATE checks SET 
                    check_number=?, amount=?, due_date=?, bank=?, status=?, description=?,
                    issuer_name=?, issuer_national_id=?, issuer_phone=?,
                    receiver_name=?, receiver_national_id=?, receiver_phone=? 
                    WHERE id=?""",
                    (
                        data["check_number"], data["amount"], data["due_date"],
                        data["bank"], data["status"], data["description"],
                        data["issuer_name"], data["issuer_national_id"], data["issuer_phone"],
                        data["receiver_name"], data["receiver_national_id"], data["receiver_phone"],
                        check_id
                    )
                )
            QMessageBox.information(self, "موفق", "چک با موفقیت ویرایش شد.")
            self._load_checks()
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ویرایش چک:\n{str(e)}")

    def _delete_check(self, check_id):
        """حذف چک"""
        reply = QMessageBox.question(
            self, "تایید حذف", "آیا از حذف این چک اطمینان دارید؟",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                with get_connection() as conn:
                    conn.execute("DELETE FROM checks WHERE id = ?", (check_id,))
                self._load_checks()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در حذف چک:\n{str(e)}")

    # --- توابع پیش‌نمایش، چاپ و خروجی ---
    def _create_checks_html(self) -> str:
        """ساخت HTML برای چاپ چک‌ها - منطبق با تصویر (راست‌چین با 12 ستون)"""
        html = """
        <html>
        <head>
        <style>
            body {
                font-family: 'Tahoma', 'Iranian Sans', sans-serif;
                direction: rtl;
                text-align: right;
                font-size: 12pt;
            }
            h2 {
                color: #4d3a78;
                text-align: center;
                font-size: 22pt;
                margin-bottom: 15px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 15px;
                table-layout: fixed;
                direction: rtl;
            }
            th {
                background-color: #7c3aed;
                color: white;
                padding: 8px;
                font-size: 11pt;
                border: 1px solid #ddd;
                text-align: center;
                vertical-align: middle;
            }
            td {
                padding: 6px;
                border-bottom: 1px solid #ddd;
                border: 1px solid #ddd;
                font-size: 10pt;
                text-align: center;
                vertical-align: middle;
            }
            .col-check { width: 10%; }
            .col-amount { width: 10%; }
            .col-date { width: 10%; }
            .col-bank { width: 8%; }
            .col-type { width: 7%; }
            .col-desc { width: 10%; }
            .col-issuer { width: 8%; }
            .col-issuer-national { width: 10%; }
            .col-issuer-phone { width: 10%; }
            .col-receiver { width: 7%; }
            .col-receiver-national { width: 10%; }
            .col-receiver-phone { width: 10%; }
        </style>
        </head>
        <body>
        <h2>لیست چک‌ها</h2>
        <table>
            <thead>
                <tr>
                    <th class="col-check">شماره چک</th>
                    <th class="col-amount">مبلغ</th>
                    <th class="col-date">تاریخ سررسید</th>
                    <th class="col-bank">بانک</th>
                    <th class="col-type">نوع</th>
                    <th class="col-desc">توضیحات</th>
                    <th class="col-issuer">صادرکننده</th>
                    <th class="col-issuer-national">شماره ملی (صادر)</th>
                    <th class="col-issuer-phone">موبایل (صادر)</th>
                    <th class="col-receiver">دریافت‌کننده</th>
                    <th class="col-receiver-national">شماره ملی (دریافت)</th>
                    <th class="col-receiver-phone">موبایل (دریافت)</th>
                </tr>
            </thead>
            <tbody>
        """
        try:
            with get_connection() as conn:
                rows = conn.execute("SELECT * FROM checks ORDER BY id DESC").fetchall()
                for row in rows:
                    html += f"""
                    <tr>
                        <td class="col-check">{row['check_number']}</td>
                        <td class="col-amount">{int(row['amount']):,} ریال</td>
                        <td class="col-date">{row['due_date']}</td>
                        <td class="col-bank">{row['bank']}</td>
                        <td class="col-type">{row['status']}</td>
                        <td class="col-desc">{row['description']}</td>
                        <td class="col-issuer">{row['issuer_name']}</td>
                        <td class="col-issuer-national">{row['issuer_national_id']}</td>
                        <td class="col-issuer-phone">{row['issuer_phone']}</td>
                        <td class="col-receiver">{row['receiver_name']}</td>
                        <td class="col-receiver-national">{row['receiver_national_id']}</td>
                        <td class="col-receiver-phone">{row['receiver_phone']}</td>
                    </tr>
                    """
        except:
            pass
        html += "</tbody></table></body></html>"
        return html

    def _preview_checks(self):
        """پیش‌نمایش چک‌ها"""
        doc = QTextDocument()
        doc.setHtml(self._create_checks_html())
        preview = QPrintPreviewDialog()
        preview.setWindowTitle("پیش‌نمایش چک‌ها")
        preview.paintRequested.connect(lambda printer: doc.print_(printer))
        preview.exec()

    def _print_checks(self):
        """چاپ چک‌ها"""
        doc = QTextDocument()
        doc.setHtml(self._create_checks_html())
        printer = QPrinter(QPrinter.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec() == QPrintDialog.Accepted:
            doc.print_(printer)

    def _export_checks_pdf(self):
        """ذخیره چک‌ها به صورت PDF و باز کردن خودکار"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره چک‌ها به صورت PDF", "checks.pdf", "PDF Files (*.pdf)"
        )
        if file_path:
            doc = QTextDocument()
            doc.setHtml(self._create_checks_html())
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            doc.print_(printer)
            QMessageBox.information(self, "موفق", "فایل PDF با موفقیت ذخیره شد.")
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))

    def _export_checks_excel(self):
        """ذخیره چک‌ها به صورت Excel"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(self, "خطا", "کتابخانه openpyxl نصب نیست. لطفاً نصب کنید.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره چک‌ها به صورت Excel", "checks.xlsx", "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "چک‌ها"
                ws.sheet_view.rightToLeft = True  # راست چین

                # هدر (منطبق با تصویر)
                headers = [
                    "شماره چک", "مبلغ", "تاریخ سررسید", "بانک", "نوع", "توضیحات",
                    "صادرکننده", "شماره ملی (صادرکننده)", "موبایل (صادرکننده)",
                    "دریافت‌کننده", "شماره ملی (دریافت‌کننده)", "موبایل (دریافت‌کننده)"
                ]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # داده‌ها
                with get_connection() as conn:
                    rows = conn.execute("SELECT * FROM checks ORDER BY id DESC").fetchall()
                    for r, row in enumerate(rows, start=2):
                        ws.cell(row=r, column=1, value=row["check_number"])
                        ws.cell(row=r, column=2, value=row["amount"])
                        ws.cell(row=r, column=3, value=row["due_date"])
                        ws.cell(row=r, column=4, value=row["bank"])
                        ws.cell(row=r, column=5, value=row["status"])
                        ws.cell(row=r, column=6, value=row["description"])
                        ws.cell(row=r, column=7, value=row["issuer_name"])
                        ws.cell(row=r, column=8, value=row["issuer_national_id"])
                        ws.cell(row=r, column=9, value=row["issuer_phone"])
                        ws.cell(row=r, column=10, value=row["receiver_name"])
                        ws.cell(row=r, column=11, value=row["receiver_national_id"])
                        ws.cell(row=r, column=12, value=row["receiver_phone"])

                wb.save(file_path)
                QMessageBox.information(self, "موفق", "فایل Excel با موفقیت ذخیره شد.")
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در ذخیره Excel:\n{str(e)}")


class AccountingPage(QWidget):
    """صفحه کامل سیستم حسابداری"""
    def __init__(self):
        super().__init__()
        self.setObjectName("accountingPage")
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        title = QLabel("📊 سیستم حسابداری")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #4d3a78;")
        layout.addWidget(title)

        self.tabs = QTabWidget()
        self.tabs.setObjectName("accountingTabs")

        # تب اسناد
        self.docs_tab = QWidget()
        docs_layout = QVBoxLayout(self.docs_tab)
        docs_layout.setContentsMargins(10, 10, 10, 10)
        docs_layout.setSpacing(10)

        toolbar_layout = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("جستجو بر اساس شماره سند یا شرح...")
        self.search_edit.setStyleSheet("padding: 7px; border-radius: 8px; border: 1px solid #cfc8ea;")
        self.search_edit.textChanged.connect(self._search_docs)
        toolbar_layout.addWidget(self.search_edit, 1)

        self.btn_new_doc = QPushButton("➕ ثبت سند جدید")
        self.btn_new_doc.setStyleSheet(
            "QPushButton{background:#7c3aed;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#6d28d9;}"
        )
        self.btn_new_doc.clicked.connect(self._create_new_doc)
        toolbar_layout.addWidget(self.btn_new_doc)

        # دکمه پیش‌نمایش
        self.preview_docs_btn = QPushButton("👁 پیش‌نمایش")
        self.preview_docs_btn.setStyleSheet(
            "QPushButton{background:#0ea5e9;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#0284c7;}"
        )
        self.preview_docs_btn.clicked.connect(self._preview_docs)
        toolbar_layout.addWidget(self.preview_docs_btn)

        # دکمه چاپ
        self.print_docs_btn = QPushButton("🖨 چاپ")
        self.print_docs_btn.setStyleSheet(
            "QPushButton{background:#16a34a;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#15803d;}"
        )
        self.print_docs_btn.clicked.connect(self._print_docs)
        toolbar_layout.addWidget(self.print_docs_btn)

        # دکمه ذخیره PDF
        self.pdf_docs_btn = QPushButton("📄 خروجی PDF")
        self.pdf_docs_btn.setStyleSheet(
            "QPushButton{background:#dc2626;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#b91c1c;}"
        )
        self.pdf_docs_btn.clicked.connect(self._export_docs_pdf)
        toolbar_layout.addWidget(self.pdf_docs_btn)

        # دکمه ذخیره Excel
        self.excel_docs_btn = QPushButton("📊 خروجی Excel")
        self.excel_docs_btn.setStyleSheet(
            "QPushButton{background:#f59e0b;color:white;border-radius:8px;padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#d97706;}"
        )
        self.excel_docs_btn.clicked.connect(self._export_docs_excel)
        toolbar_layout.addWidget(self.excel_docs_btn)

        docs_layout.addLayout(toolbar_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["شماره سند", "تاریخ", "شرح", "بدهکار", "بستانکار", "", ""])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        docs_layout.addWidget(self.table)

        self.tabs.addTab(self.docs_tab, "📒 اسناد حسابداری")

        # تب چک‌ها
        self.checks_tab = ChecksTab()
        self.tabs.addTab(self.checks_tab, "🏦 ثبت چک")

        layout.addWidget(self.tabs)
        self._load_data()

    def _load_data(self):
        try:
            with get_connection() as conn:
                rows = conn.execute("SELECT * FROM accounting_docs ORDER BY id DESC").fetchall()
                self._populate_table(rows)
        except Exception:
            self.table.setRowCount(0)

    def _populate_table(self, rows):
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            values = [
                str(row["doc_number"]),
                str(row["date"]),
                str(row["description"]),
                f"{int(row['debit']):,} ریال",
                f"{int(row['credit']):,} ریال"
            ]
            for c, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(r, c, item)

            edit_btn = QPushButton("✏️ ویرایش")
            edit_btn.setStyleSheet(
                "QPushButton{background:#0ea5e9;color:white;border-radius:6px;padding:4px 10px;font-weight:bold;border:none;}"
                "QPushButton:hover{background:#0284c7;}"
            )
            row_dict = dict(row)
            edit_btn.clicked.connect(lambda _, d=row_dict: self._edit_doc(d))
            self.table.setCellWidget(r, 5, edit_btn)

            delete_btn = QPushButton("🗑️ حذف")
            delete_btn.setStyleSheet(
                "QPushButton{background:#ef4444;color:white;border-radius:6px;padding:4px 10px;font-weight:bold;border:none;}"
                "QPushButton:hover{background:#dc2626;}"
            )
            doc_id = row["id"]
            delete_btn.clicked.connect(lambda _, did=doc_id: self._delete_doc(did))
            self.table.setCellWidget(r, 6, delete_btn)

    def _search_docs(self, text):
        term = f"%{text.strip()}%"
        try:
            with get_connection() as conn:
                rows = conn.execute(
                    "SELECT * FROM accounting_docs WHERE doc_number LIKE ? OR description LIKE ? ORDER BY id DESC",
                    (term, term)
                ).fetchall()
                self._populate_table(rows)
        except Exception:
            pass

    def _create_new_doc(self):
        dlg = NewDocDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        data = dlg.get_data()
        try:
            with get_connection() as conn:
                conn.execute(
                    "INSERT INTO accounting_docs (doc_number, date, description, debit, credit) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (data["doc_number"], data["date"], data["description"], data["debit"], data["credit"])
                )
            QMessageBox.information(self, "موفق", "سند حسابداری با موفقیت ثبت شد.")
            self._load_data()
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ثبت سند:\n{str(e)}")

    def _edit_doc(self, row_data):
        dlg = NewDocDialog(self, doc_data=row_data)
        if dlg.exec() != QDialog.Accepted:
            return
        data = dlg.get_data()
        try:
            with get_connection() as conn:
                conn.execute(
                    "UPDATE accounting_docs SET doc_number=?, date=?, description=?, debit=?, credit=? WHERE id=?",
                    (data["doc_number"], data["date"], data["description"], data["debit"], data["credit"], row_data["id"])
                )
            QMessageBox.information(self, "موفق", "سند حسابداری ویرایش شد.")
            self._load_data()
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ویرایش سند:\n{str(e)}")

    def _delete_doc(self, doc_id):
        reply = QMessageBox.question(
            self, "تایید حذف", "آیا از حذف این سند حسابداری اطمینان دارید؟",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                with get_connection() as conn:
                    conn.execute("DELETE FROM accounting_docs WHERE id = ?", (doc_id,))
                self._load_data()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در حذف سند:\n{str(e)}")

    # --- توابع پیش‌نمایش، چاپ و خروجی ---
    def _create_docs_html(self) -> str:
        """ساخت HTML برای چاپ اسناد - دقیقاً منطبق با تصویر (راست‌چین)"""
        html = """
        <html>
        <head>
        <style>
            body {
                font-family: 'Tahoma', 'Iranian Sans', sans-serif;
                direction: rtl;
                text-align: right;
                font-size: 12pt;
            }
            h2 {
                color: #4d3a78;
                text-align: center;
                font-size: 22pt;
                margin-bottom: 15px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 15px;
                table-layout: fixed;
                direction: rtl;
            }
            th {
                background-color: #7c3aed;
                color: white;
                padding: 8px;
                font-size: 11pt;
                border: 1px solid #ddd;
                text-align: center;
                vertical-align: middle;
            }
            td {
                padding: 6px;
                border-bottom: 1px solid #ddd;
                border: 1px solid #ddd;
                font-size: 10pt;
                text-align: center;
                vertical-align: middle;
            }
            .col-num { width: 15%; }
            .col-date { width: 15%; }
            .col-desc { width: 30%; }
            .col-debit { width: 20%; }
            .col-credit { width: 20%; }
        </style>
        </head>
        <body>
        <h2>لیست اسناد حسابداری</h2>
        <table>
            <thead>
                <tr>
                    <th class="col-num">شماره سند</th>
                    <th class="col-date">تاریخ</th>
                    <th class="col-desc">شرح</th>
                    <th class="col-debit">بدهکار</th>
                    <th class="col-credit">بستانکار</th>
                </tr>
            </thead>
            <tbody>
        """
        try:
            with get_connection() as conn:
                rows = conn.execute("SELECT * FROM accounting_docs ORDER BY id DESC").fetchall()
                for row in rows:
                    html += f"""
                    <tr>
                        <td class="col-num">{row['doc_number']}</td>
                        <td class="col-date">{row['date']}</td>
                        <td class="col-desc">{row['description']}</td>
                        <td class="col-debit">{int(row['debit']):,} ریال</td>
                        <td class="col-credit">{int(row['credit']):,} ریال</td>
                    </tr>
                    """
        except:
            pass
        html += "</tbody></table></body></html>"
        return html

    def _preview_docs(self):
        """پیش‌نمایش اسناد"""
        doc = QTextDocument()
        doc.setHtml(self._create_docs_html())
        preview = QPrintPreviewDialog()
        preview.setWindowTitle("پیش‌نمایش اسناد حسابداری")
        preview.paintRequested.connect(lambda printer: doc.print_(printer))
        preview.exec()

    def _print_docs(self):
        """چاپ اسناد"""
        doc = QTextDocument()
        doc.setHtml(self._create_docs_html())
        printer = QPrinter(QPrinter.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec() == QPrintDialog.Accepted:
            doc.print_(printer)

    def _export_docs_pdf(self):
        """ذخیره اسناد به صورت PDF و باز کردن خودکار"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره اسناد به صورت PDF", "accounting_docs.pdf", "PDF Files (*.pdf)"
        )
        if file_path:
            doc = QTextDocument()
            doc.setHtml(self._create_docs_html())
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            doc.print_(printer)
            QMessageBox.information(self, "موفق", "فایل PDF با موفقیت ذخیره شد.")
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))

    def _export_docs_excel(self):
        """ذخیره اسناد به صورت Excel"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(self, "خطا", "کتابخانه openpyxl نصب نیست. لطفاً نصب کنید.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره اسناد به صورت Excel", "accounting_docs.xlsx", "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "اسناد حسابداری"
                ws.sheet_view.rightToLeft = True  # راست چین

                # هدر (منطبق با تصویر)
                headers = ["شماره سند", "تاریخ", "شرح", "بدهکار", "بستانکار"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # داده‌ها
                with get_connection() as conn:
                    rows = conn.execute("SELECT * FROM accounting_docs ORDER BY id DESC").fetchall()
                    for r, row in enumerate(rows, start=2):
                        ws.cell(row=r, column=1, value=row["doc_number"])
                        ws.cell(row=r, column=2, value=row["date"])
                        ws.cell(row=r, column=3, value=row["description"])
                        ws.cell(row=r, column=4, value=row["debit"])
                        ws.cell(row=r, column=5, value=row["credit"])

                wb.save(file_path)
                QMessageBox.information(self, "موفق", "فایل Excel با موفقیت ذخیره شد.")
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در ذخیره Excel:\n{str(e)}")